"""Excel output generation for the OpenFloat Data Formatter.

Writes a two-sheet .xlsx file matching the OpenFloat Transactions Template:
- Accounts: transformed data rows
- Allowed Types: verbatim copy from the reference template

The Allowed Types sheet must be copied exactly (including trailing spaces
like "SPA NAKURU RURAL ").
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd

from .config import OPENFLOAT_ACCOUNTS_COLUMNS
from .models import OutputRow


def load_allowed_types(template_path: str | Path) -> list[str]:
    """Load the Allowed Types list from the OpenFloat reference template.

    Reads column A of the 'Allowed Types' sheet, skipping the header row.
    Values are returned as-is to preserve exact strings (including trailing
    spaces like "SPA NAKURU RURAL ").

    Args:
        template_path: Path to the reference template .xlsx file.

    Returns:
        A list of allowed type strings.

    Raises:
        FileNotFoundError: If the template file does not exist.
        KeyError: If the 'Allowed Types' sheet is missing.
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(str(template_path), read_only=True, data_only=True)
    if "Allowed Types" not in wb.sheetnames:
        wb.close()
        raise KeyError(
            f"'Allowed Types' sheet not found in template. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb["Allowed Types"]
    types: list[str] = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        value = row[0]
        if value is not None:
            # Preserve exact string including trailing spaces
            types.append(str(value))

    wb.close()
    return types


def write_openfloat_excel(
    rows: list[OutputRow],
    allowed_types: list[str],
    output_path: str | Path | None = None,
) -> BytesIO | Path:
    """Write a two-sheet OpenFloat-ready Excel file.

    Args:
        rows: List of OutputRow objects for the Accounts sheet.
        allowed_types: List of allowed type strings for the Allowed Types sheet.
        output_path: If provided, writes to this file path. If None, returns BytesIO.

    Returns:
        BytesIO containing the .xlsx if output_path is None,
        otherwise the Path of the written file.
    """
    # Create workbook
    wb = openpyxl.Workbook()

    # --- Accounts sheet ---
    ws_accounts = wb.active
    ws_accounts.title = "Accounts"

    # Write header row
    ws_accounts.append(OPENFLOAT_ACCOUNTS_COLUMNS)

    # Write data rows
    for row in rows:
        ws_accounts.append([
            row.account_type,
            row.account_name,
            row.account_number,
            row.till_or_paybill_number,
            row.till_or_paybill_business_name,
            row.notification_phone_number,
            row.amount,
            row.remark,
        ])

    # --- Allowed Types sheet ---
    ws_types = wb.create_sheet(title="Allowed Types")

    # Write each type as a single-cell row (column A only)
    for type_name in allowed_types:
        ws_types.append([type_name])

    # Save to BytesIO or file
    if output_path is not None:
        output_path = Path(output_path)
        wb.save(str(output_path))
        wb.close()
        return output_path
    else:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        wb.close()
        return buffer