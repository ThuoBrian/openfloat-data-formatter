"""Tests for the transformer module — end-to-end transformation pipeline."""

import pandas as pd
import pytest

from backend.config import Settings
from backend.transformer import transform, _build_output_rows


class TestTransformWithSampleData:
    """End-to-end transformation with the real sample CSV."""

    def test_transform_sample_csv(self, sample_csv_path, default_config):
        """Transform the sample CSV and verify output."""
        result = transform(sample_csv_path, default_config)
        assert result.output is not None
        assert result.output_row_count > 0

    def test_transform_output_row_count(self, sample_csv_path, default_config):
        """All 196 rows in the sample CSV should transform successfully."""
        result = transform(sample_csv_path, default_config)
        # All rows have consent=Yes, valid phones, valid amounts, known networks
        assert result.output_row_count == 196
        assert result.validation_report.total_rows == 196

    def test_transform_no_errors(self, sample_csv_path, default_config):
        """No hard errors in the sample data."""
        result = transform(sample_csv_path, default_config)
        assert len(result.validation_report.errors) == 0


class TestTransformOutputFormat:
    """Verify the output Excel file format."""

    def test_output_is_bytes_io(self, sample_csv_path, default_config):
        """Output is a BytesIO object when no output_path is specified."""
        from io import BytesIO

        result = transform(sample_csv_path, default_config)
        assert isinstance(result.output, BytesIO)

    def test_output_readable_as_excel(self, sample_csv_path, default_config):
        """Output can be read as a valid Excel file."""
        import openpyxl

        result = transform(sample_csv_path, default_config)
        result.output.seek(0)
        wb = openpyxl.load_workbook(result.output)
        assert "Accounts" in wb.sheetnames
        assert "Allowed Types" in wb.sheetnames
        wb.close()


class TestBuildOutputRows:
    """Test the _build_output_rows helper."""

    def test_basic_transformation(self, minimal_df, default_config):
        """Minimal valid DataFrame produces correct output rows."""
        rows, errors = _build_output_rows(minimal_df, default_config)
        assert len(rows) == 2
        assert len(errors) == 0

    def test_phone_normalization(self, minimal_df, default_config):
        """Phone numbers are normalized with 254 prefix."""
        rows, _ = _build_output_rows(minimal_df, default_config)
        assert rows[0].account_number == "254712345678"
        assert rows[0].notification_phone_number == "254712345678"

    def test_network_mapping(self, minimal_df, default_config):
        """Networks are mapped to correct account types."""
        rows, _ = _build_output_rows(minimal_df, default_config)
        assert rows[0].account_type == "Safaricom Prepaid"
        assert rows[1].account_type == "Airtel Prepaid"

    def test_consent_filter(self, minimal_df, default_config):
        """Rows with consent != Yes are excluded."""
        minimal_df.loc[0, "consent"] = "No"
        rows, errors = _build_output_rows(minimal_df, default_config)
        assert len(rows) == 1
        assert 0 in errors

    def test_remark_format(self, minimal_df, default_config):
        """Remark is formatted as 'project_name - Project_Activity'."""
        rows, _ = _build_output_rows(minimal_df, default_config)
        assert rows[0].remark == "Test Project - g05|Testing"

    def test_amount_coercion(self, minimal_df, default_config):
        """Amounts are coerced to float."""
        minimal_df["amount"] = minimal_df["amount"].astype(object)
        minimal_df.loc[0, "amount"] = "300"
        rows, _ = _build_output_rows(minimal_df, default_config)
        assert rows[0].amount == 300.0

    def test_invalid_phone_excluded(self, minimal_df, default_config):
        """Rows with invalid phone numbers are excluded."""
        minimal_df.loc[0, "airtime_phone"] = "123"  # Too short
        rows, errors = _build_output_rows(minimal_df, default_config)
        assert len(rows) == 1
        assert 0 in errors

    def test_unmapped_network_excluded(self, minimal_df, default_config):
        """Rows with unmapped networks are excluded."""
        minimal_df.loc[0, "network"] = "Orange"
        rows, errors = _build_output_rows(minimal_df, default_config)
        assert len(rows) == 1
        assert 0 in errors

    def test_invalid_amount_excluded(self, minimal_df, default_config):
        """Rows with invalid amounts are excluded."""
        minimal_df["amount"] = minimal_df["amount"].astype(object)
        minimal_df.loc[0, "amount"] = "abc"
        rows, errors = _build_output_rows(minimal_df, default_config)
        assert len(rows) == 1
        assert 0 in errors

    def test_optional_fields_blank(self, minimal_df, default_config):
        """Till/Paybill fields are blank in output."""
        rows, _ = _build_output_rows(minimal_df, default_config)
        assert rows[0].till_or_paybill_number == ""
        assert rows[0].till_or_paybill_business_name == ""