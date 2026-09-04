"""Tests for the statement module — OpenFloat Transaction Statement reporting.

All statement workbooks are SYNTHETIC (built by the make_statement_workbook
fixture). The real exports in sample_report_output/ contain personal data
and are not tracked by git — the test suite must never reference them.
"""

from datetime import datetime

import pandas as pd
import pytest

from backend.statement import (
    build_statement_report,
    is_successful_status,
    parse_statement_date,
    parse_statement_file,
    reconcile,
    rollup_by_case,
    summarize_transactions,
)

GOOD_DATE = "24/08/2026 02:56:42 PM"
GOOD_REMARK = "C#37154 13054AF RESP AIRTIME-KSH27900 d05"


def _txn(
    status="Successful",
    phone=254712345678,
    amount=100,
    remark=GOOD_REMARK,
    **overrides,
):
    """A default statement data row; keyword args override any header column."""
    row = {
        "Approval Id": 14886185,
        "Transaction Id": 18488654,
        "Transaction Type": "Payment",
        "Transaction Status": status,
        "Date": GOOD_DATE,
        "Account Name": 9019830,
        "Account Number": phone,
        "Account Type": "Partner",
        "Remark": remark,
        "Initiated By": "Test User",
        "Approved/Rejected By": "Test Approver",
        "Amount": amount,
    }
    row.update(overrides)
    return row


class TestParseStatementDate:
    """Test statement Date cell parsing."""

    def test_parses_expected_format(self):
        """'DD/MM/YYYY hh:mm:ss AM/PM' strings parse to datetimes."""
        date, error = parse_statement_date(GOOD_DATE)
        assert error is None
        assert date == datetime(2026, 8, 24, 14, 56, 42)

    def test_empty_is_not_an_error(self):
        """None / blank cells parse to (None, None) — nothing to parse."""
        assert parse_statement_date(None) == (None, None)
        assert parse_statement_date("") == (None, None)
        assert parse_statement_date("   ") == (None, None)

    def test_bad_string_soft_error(self):
        """Unparseable strings return an error, not an exception."""
        date, error = parse_statement_date("not a date")
        assert date is None
        assert error is not None

    def test_excel_serial(self):
        """Excel serial numbers convert via the 1899-12-30 epoch."""
        date, error = parse_statement_date(45000)
        assert error is None
        assert date == datetime(2023, 3, 15)

    def test_datetime_passthrough(self):
        """Already-parsed datetimes pass through unchanged."""
        now = datetime(2026, 8, 24, 10, 0, 0)
        assert parse_statement_date(now) == (now, None)


class TestIsSuccessfulStatus:
    """Test the 'paid' classification rule."""

    def test_successful_variants(self):
        """'Successful' matches case-insensitively with surrounding whitespace."""
        assert is_successful_status("Successful")
        assert is_successful_status(" SUCCESSFUL ")
        assert is_successful_status("successful")

    def test_everything_else_is_unsuccessful(self):
        """Reversed, Failed, Pending, empty, and unknown statuses are not paid."""
        for status in ("Reversed", "Failed", "Rejected", "Pending", "", None):
            assert not is_successful_status(status)


class TestParseStatementFile:
    """Test parsing of synthetic statement workbooks."""

    def test_parses_basic_successful_rows(self, make_statement_workbook):
        """Successful rows parse with amounts, normalized phones, parsed remarks."""
        buffer = make_statement_workbook(
            rows=[_txn(), _txn(phone=254798765432, amount=200)],
            footer_total=300,
        )
        transactions, footer_total, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert len(transactions) == 2
        first = transactions[0]
        assert first.is_successful
        assert first.amount == 100.0
        assert first.account_number == "254712345678"
        assert first.account_name == "9019830"
        assert first.remark_parts is not None
        assert first.remark_parts.case_number == "37154"
        assert first.date == datetime(2026, 8, 24, 14, 56, 42)

    def test_header_without_reference_id(self, make_statement_workbook):
        """12-column exports (no 'Reference Id') parse; reference_id stays empty."""
        buffer = make_statement_workbook(rows=[_txn()], include_reference_id=False)
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert transactions[0].reference_id == ""

    def test_header_with_reference_id_on_reversed_row(self, make_statement_workbook):
        """13-column exports capture the Reference Id on Reversed rows."""
        buffer = make_statement_workbook(
            rows=[_txn(status="Reversed", amount=None, **{"Reference Id": 18492036})]
        )
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert transactions[0].status == "Reversed"
        assert transactions[0].reference_id == "18492036"

    def test_footer_row_captured_not_counted(self, make_statement_workbook):
        """The grand-total footer row becomes footer_total, not a transaction."""
        buffer = make_statement_workbook(rows=[_txn()], footer_total=100)
        transactions, footer_total, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert len(transactions) == 1
        assert footer_total == 100.0

    def test_blank_rows_skipped(self, make_statement_workbook):
        """Fully blank rows in the middle of the sheet are skipped."""
        from openpyxl import load_workbook

        buffer = make_statement_workbook(rows=[_txn()], footer_total=100)
        # Insert a fully blank row after the first data row
        workbook = load_workbook(buffer)
        workbook["Transaction Statement"].append([None] * 12)
        workbook["Transaction Statement"].append([None] * 12 + [100])
        buffer.seek(0)

        transactions, footer_total, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert len(transactions) == 1
        assert footer_total == 100.0

    def test_reversed_row_is_none_amount_and_unsuccessful(self, make_statement_workbook):
        """Reversed rows keep amount=None and classify as unsuccessful."""
        buffer = make_statement_workbook(
            rows=[_txn(), _txn(status="Reversed", amount=None, phone=254722345678)]
        )
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        reversed_row = transactions[1]
        assert reversed_row.amount is None
        assert not reversed_row.is_successful

    def test_account_name_string_variant(self, make_statement_workbook):
        """String account ids like 'I220008' are kept verbatim."""
        buffer = make_statement_workbook(rows=[_txn(**{"Account Name": "I220008"})])
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert transactions[0].account_name == "I220008"

    def test_unknown_status_counted_not_crash(self, make_statement_workbook):
        """Unknown statuses classify as unsuccessful without raising."""
        buffer = make_statement_workbook(rows=[_txn(status="Failed", amount=None)])
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert transactions[0].status == "Failed"
        assert not transactions[0].is_successful

    def test_missing_required_column_returns_error(self, make_statement_workbook):
        """A sheet missing a required column yields an error and no transactions."""
        from io import BytesIO

        from openpyxl import load_workbook

        buffer = make_statement_workbook(rows=[])
        workbook = load_workbook(buffer)
        worksheet = workbook["Transaction Statement"]
        # Header order: Approval Id, Transaction Id, Transaction Type,
        # Transaction Status, Date, Account Name, Account Number, Account
        # Type, Remark, ... — column 9 is 'Remark' (1-based, openpyxl).
        worksheet.delete_cols(9)
        modified = BytesIO()
        workbook.save(modified)
        modified.seek(0)

        transactions, footer_total, warnings, errors = parse_statement_file(modified)
        assert transactions == []
        assert len(errors) == 1
        assert "Remark" in errors[0]

    def test_missing_sheet_returns_error(self, make_statement_workbook):
        """A workbook without the 'Transaction Statement' sheet yields an error."""
        buffer = make_statement_workbook(rows=[], sheet_name="Wrong Sheet")
        transactions, footer_total, warnings, errors = parse_statement_file(buffer)
        assert transactions == []
        assert len(errors) == 1
        assert "Transaction Statement" in errors[0]

    def test_bad_date_soft_warning_date_raw_kept(self, make_statement_workbook):
        """An unparseable date warns and keeps the raw cell text."""
        buffer = make_statement_workbook(rows=[_txn(Date="not a date")])
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert transactions[0].date is None
        assert transactions[0].date_raw == "not a date"
        assert any("Date" in warning for warning in warnings)

    def test_unparseable_remark_soft_warning(self, make_statement_workbook):
        """A remark that doesn't match the case format warns; remark text is kept."""
        buffer = make_statement_workbook(rows=[_txn(remark="garbage")])
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert transactions[0].remark_parts is None
        assert transactions[0].remark == "garbage"
        assert any("case_remark" in warning for warning in warnings)

    def test_unnormalizable_account_number_warning_raw_kept(self, make_statement_workbook):
        """A non-9-digit account number warns and keeps the raw value."""
        buffer = make_statement_workbook(rows=[_txn(**{"Account Number": 12345})])
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert errors == []
        assert transactions[0].account_number == "12345"
        assert transactions[0].account_number_error is not None
        assert any("Phone number" in warning for warning in warnings)

    def test_local_phone_input_normalizes_to_statement_key(self, make_statement_workbook):
        """An Account Number stored with local prefix still normalizes to 254XXXXXXXXX."""
        buffer = make_statement_workbook(rows=[_txn(**{"Account Number": "0712345678"})])
        transactions, _, warnings, errors = parse_statement_file(buffer)
        assert transactions[0].account_number == "254712345678"


class TestSummarize:
    """Test summarize_transactions aggregation."""

    def _transactions(self, make_statement_workbook):
        buffer = make_statement_workbook(
            rows=[
                _txn(amount=100),
                _txn(phone=254798765432, amount=200),
                _txn(status="Reversed", amount=None, phone=254722345678),
                _txn(status="Failed", amount=None, phone=254732345678),
            ]
        )
        transactions, _, _, errors = parse_statement_file(buffer)
        assert errors == []
        return transactions

    def test_counts_by_status_histogram(self, make_statement_workbook):
        """counts_by_status records every status seen."""
        summary = summarize_transactions(self._transactions(make_statement_workbook), "f.xlsx")
        assert summary.counts_by_status == {"Successful": 2, "Reversed": 1, "Failed": 1}

    def test_total_disbursed_sums_successful_only(self, make_statement_workbook):
        """total_disbursed sums Successful amounts; unsuccessful rows contribute 0."""
        summary = summarize_transactions(self._transactions(make_statement_workbook), "f.xlsx")
        assert summary.total_rows == 4
        assert summary.successful_count == 2
        assert summary.unsuccessful_count == 2
        assert summary.total_disbursed == 300.0

    def test_success_rate(self, make_statement_workbook):
        """success_rate is successful / total."""
        summary = summarize_transactions(self._transactions(make_statement_workbook), "f.xlsx")
        assert summary.success_rate == 0.5

    def test_footer_total_mismatch_flag(self, make_statement_workbook):
        """A footer that disagrees with the computed total sets footer_matches False."""
        summary = summarize_transactions(
            self._transactions(make_statement_workbook), "f.xlsx", footer_total=999
        )
        assert summary.footer_matches is False

    def test_footer_matches_when_equal(self, make_statement_workbook):
        """A footer equal to the computed total sets footer_matches True."""
        summary = summarize_transactions(
            self._transactions(make_statement_workbook), "f.xlsx", footer_total=300
        )
        assert summary.footer_matches is True

    def test_no_footer_footer_matches_none(self, make_statement_workbook):
        """Without a footer, footer_matches stays None (unknown, not False)."""
        summary = summarize_transactions(self._transactions(make_statement_workbook), "f.xlsx")
        assert summary.footer_total is None
        assert summary.footer_matches is None

    def test_empty_transactions(self):
        """An empty summary reports zero rows and a 0.0 success rate."""
        summary = summarize_transactions([], "f.xlsx")
        assert summary.total_rows == 0
        assert summary.success_rate == 0.0


class TestCaseRollup:
    """Test per-case aggregation from parsed remarks."""

    def test_groups_by_case(self, make_statement_workbook):
        """Rows for the same case group into one rollup."""
        buffer = make_statement_workbook(rows=[_txn(), _txn(phone=254798765432)])
        transactions, _, _, _ = parse_statement_file(buffer)
        rollups, unparsed = rollup_by_case(transactions)
        assert unparsed == 0
        assert len(rollups) == 1
        assert rollups[0].total_rows == 2
        assert rollups[0].successful_count == 2
        assert rollups[0].disbursed_total == 200.0
        assert rollups[0].remark_amount == 27900.0

    def test_difference_vs_remark_amount(self, make_statement_workbook):
        """difference = disbursed_total - remark_amount (the shortfall flag)."""
        buffer = make_statement_workbook(
            rows=[_txn(amount=100), _txn(phone=254798765432, amount=200)]
        )
        transactions, _, _, _ = parse_statement_file(buffer)
        rollups, _ = rollup_by_case(transactions)
        assert rollups[0].difference == 300.0 - 27900.0

    def test_separate_cases_sorted(self, make_statement_workbook):
        """Distinct cases produce distinct rollups, sorted by case number."""
        buffer = make_statement_workbook(
            rows=[
                _txn(remark="C#37200 13054AF RESP AIRTIME-KSH100 d05"),
                _txn(remark="C#37154 13054AF RESP AIRTIME-KSH100 d05"),
            ]
        )
        transactions, _, _, _ = parse_statement_file(buffer)
        rollups, _ = rollup_by_case(transactions)
        assert [rollup.case_number for rollup in rollups] == ["37154", "37200"]

    def test_unparsed_remarks_excluded_and_counted(self, make_statement_workbook):
        """Rows with unparseable remarks are excluded and counted as unparsed."""
        buffer = make_statement_workbook(
            rows=[_txn(), _txn(remark="garbage")]
        )
        transactions, _, _, _ = parse_statement_file(buffer)
        rollups, unparsed = rollup_by_case(transactions)
        assert unparsed == 1
        assert rollups[0].total_rows == 1


class TestBuildStatementReport:
    """Test the orchestrator across multiple statement files."""

    def test_multiple_files_combined(self, make_statement_workbook):
        """Totals combine across files; per-file summaries stay separate."""
        file_a = make_statement_workbook(
            rows=[_txn(amount=100)], footer_total=100, include_reference_id=True
        )
        file_b = make_statement_workbook(
            rows=[_txn(phone=254798765432, amount=200)],
            footer_total=200,
            include_reference_id=False,
        )
        report = build_statement_report(
            [file_a, file_b], source_names=["a.xlsx", "b.xlsx"]
        )
        assert report.errors == []
        assert len(report.file_summaries) == 2
        assert report.combined.total_rows == 2
        assert report.combined.successful_count == 2
        assert report.combined.total_disbursed == 300.0
        assert report.reconciliation is None

    def test_one_bad_file_does_not_kill_others(self, make_statement_workbook):
        """A file with a structural error still lets the good file report."""
        good = make_statement_workbook(rows=[_txn()], footer_total=100)
        bad = make_statement_workbook(rows=[], sheet_name="Wrong Sheet")
        report = build_statement_report([good, bad], source_names=["good.xlsx", "bad.xlsx"])
        assert len(report.errors) == 1
        assert "bad.xlsx" in report.errors[0]
        assert report.combined.total_rows == 1

    def test_unreadable_file_reports_error(self):
        """A corrupt (non-xlsx) buffer yields an error entry, not a crash."""
        from io import BytesIO

        report = build_statement_report(
            [BytesIO(b"not an excel file")], source_names=["corrupt.xlsx"]
        )
        assert len(report.errors) == 1
        assert "corrupt.xlsx" in report.errors[0]
        assert report.combined.total_rows == 0


class TestReconcile:
    """Test reconciliation between a Process Maker input and statements."""

    @pytest.fixture
    def statement_transactions(self, make_statement_workbook):
        """Statement rows matching the pm_input_df fixture's phones.

        - 254712345678: Successful 100 (input row amount 100 → clean)
        - 254722345678: Reversed only            → matched_not_paid
        - 254742345678: Successful 50             → amount mismatch vs input 100
        - 254752345678: two Successful of 100     → matches input sum 200, multiply paid
        """
        buffer = make_statement_workbook(
            rows=[
                _txn(phone=254712345678, amount=100),
                _txn(phone=254722345678, status="Reversed", amount=None),
                _txn(phone=254742345678, amount=50),
                _txn(phone=254752345678, amount=100),
                _txn(phone=254752345678, amount=100, **{"Transaction Id": 999}),
            ]
        )
        transactions, _, _, errors = parse_statement_file(buffer)
        assert errors == []
        return transactions

    def test_all_buckets(self, pm_input_df, statement_transactions):
        """Every reconciliation bucket is populated as documented in the fixture."""
        result = reconcile(pm_input_df, statement_transactions)
        assert result.input_rows == 6

        paid_phones = [entry.phone for entry in result.matched_paid]
        assert "254712345678" in paid_phones
        assert "254742345678" in paid_phones  # paid, but amount-mismatched

        not_paid = [entry.phone for entry in result.matched_not_paid]
        assert not_paid == ["254722345678"]

        missing = [entry.phone for entry in result.missing_from_statement]
        assert missing == ["254733345678"]

        assert result.duplicate_input_phones == ["254752345678"]
        assert result.multiply_paid_phones == ["254752345678"]
        assert result.statement_not_in_input == []

    def test_matched_paid_entry_details(self, pm_input_df, statement_transactions):
        """A clean matched-paid entry carries the input rows and statement totals."""
        result = reconcile(pm_input_df, statement_transactions)
        entry = next(
            e for e in result.matched_paid if e.phone == "254712345678"
        )
        assert entry.input_row_numbers == [2]
        assert entry.input_amount == 100.0
        assert entry.successful_count == 1
        assert entry.successful_total == 100.0
        assert entry.notes == []

    def test_amount_mismatch_note(self, pm_input_df, statement_transactions):
        """An input amount that disagrees with the statement gets a note."""
        result = reconcile(pm_input_df, statement_transactions)
        entry = next(
            e for e in result.matched_paid if e.phone == "254742345678"
        )
        assert any("differs" in note for note in entry.notes)

    def test_paid_phone_with_unsuccessful_rows_noted(self, pm_input_df, statement_transactions):
        """A phone with both Successful and unsuccessful rows lands in matched_paid with a note."""
        extra = list(statement_transactions) + [
            statement_transactions[0].model_copy(
                update={
                    "status": "Reversed",
                    "is_successful": False,
                    "amount": None,
                    "account_number": "254712345678",
                }
            )
        ]
        result = reconcile(pm_input_df, extra)
        entry = next(e for e in result.matched_paid if e.phone == "254712345678")
        assert any("unsuccessful" in note for note in entry.notes)

    def test_statement_not_in_input(self, pm_input_df, statement_transactions):
        """Statement phones absent from the input are listed with their statement counts."""
        input_df = pm_input_df[pm_input_df["airtime_phone"] == "712345678"]
        result = reconcile(input_df, statement_transactions)
        not_in_input = {
            entry.phone for entry in result.statement_not_in_input
        }
        assert not_in_input == {"254722345678", "254742345678", "254752345678"}
        entry = next(
            e
            for e in result.statement_not_in_input
            if e.phone == "254742345678"
        )
        assert entry.successful_count == 1
        assert entry.successful_total == 50.0

    def test_local_format_input_phone_matches(self, statement_transactions):
        """An input phone in local format ('07...') matches the statement's 254 number."""
        input_df = pd.DataFrame({"airtime_phone": ["0712345678"], "amount": [100]})
        result = reconcile(input_df, statement_transactions)
        assert [entry.phone for entry in result.matched_paid] == ["254712345678"]

    def test_2540_edge_case_round_trips(self, make_statement_workbook):
        """The 2540-prefixed edge case matches across both sides.

        The statement stores 254012345678; the input has the same number as
        012345678 (9 digits, leading zero kept per the normalizer's edge case).
        """
        buffer = make_statement_workbook(
            rows=[_txn(phone=254012345678, amount=100)]
        )
        transactions, _, _, errors = parse_statement_file(buffer)
        assert errors == []
        assert transactions[0].account_number == "254012345678"

        input_df = pd.DataFrame({"airtime_phone": ["012345678"], "amount": [100]})
        result = reconcile(input_df, transactions)
        assert [entry.phone for entry in result.matched_paid] == ["254012345678"]
        assert result.matched_paid[0].notes == []

    def test_unnormalizable_input_phone_noted(self, statement_transactions):
        """An input phone that can't normalize lands in missing with the error in notes."""
        input_df = pd.DataFrame({"airtime_phone": ["123"], "amount": [100]})
        result = reconcile(input_df, statement_transactions)
        assert len(result.missing_from_statement) == 1
        entry = result.missing_from_statement[0]
        assert entry.phone == "123"
        assert entry.notes and "not 9 digits" in entry.notes[0]

    def test_empty_input_df(self, statement_transactions):
        """An empty input yields empty buckets and lists all statement phones as not-in-input."""
        input_df = pd.DataFrame(columns=["airtime_phone", "amount"])
        result = reconcile(input_df, statement_transactions)
        assert result.input_rows == 0
        assert result.matched_paid == []
        assert len(result.statement_not_in_input) == 4