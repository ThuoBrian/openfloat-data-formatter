"""Tests for the writer module — Excel output generation."""

import openpyxl
import pytest

from backend.config import DEFAULT_TEMPLATE_PATH
from backend.models import OutputRow
from backend.writer import load_allowed_types, write_openfloat_excel


class TestLoadAllowedTypes:
    """Test loading Allowed Types from the reference template."""

    def test_loads_from_template(self, template_path):
        """Allowed Types are loaded from the reference template."""
        types = load_allowed_types(template_path)
        assert len(types) > 0
        assert "Safaricom Prepaid" in types
        assert "Airtel Prepaid" in types

    def test_verbatim_preservation(self, template_path):
        """The 'SPA NAKURU RURAL ' entry preserves its trailing space."""
        types = load_allowed_types(template_path)
        # Find the entry that starts with "SPA NAKURU"
        nakuru_entries = [t for t in types if "NAKURU" in t]
        assert len(nakuru_entries) == 1
        assert nakuru_entries[0] == "SPA NAKURU RURAL "  # Note trailing space

    def test_file_not_found(self):
        """Missing template file raises FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            load_allowed_types("/nonexistent/path.xlsx")

    def test_expected_type_count(self, template_path):
        """The template contains exactly 63 allowed types (header row 'Mpesa' is data, not a header)."""
        types = load_allowed_types(template_path)
        assert len(types) == 63


class TestWriteOpenfloatExcel:
    """Test Excel file generation."""

    @pytest.fixture
    def sample_output_rows(self):
        """Sample OutputRow list for testing."""
        return [
            OutputRow(
                **{
                    "Account Type": "Safaricom Prepaid",
                    "Account Name": "TEST001",
                    "Account Number": "254712345678",
                    "Till or Paybill Number": "",
                    "Till or Paybill Business Name": "",
                    "Notification Phone Number": "254712345678",
                    "Amount": 150.0,
                    "Remark": "Test Project - g05|Testing",
                }
            ),
            OutputRow(
                **{
                    "Account Type": "Airtel Prepaid",
                    "Account Name": "TEST002",
                    "Account Number": "254798765432",
                    "Till or Paybill Number": "",
                    "Till or Paybill Business Name": "",
                    "Notification Phone Number": "254798765432",
                    "Amount": 200.0,
                    "Remark": "Test Project - g05|Testing",
                }
            ),
        ]

    @pytest.fixture
    def allowed_types(self, template_path):
        """Load allowed types from the reference template."""
        return load_allowed_types(template_path)

    def test_output_has_two_sheets(self, sample_output_rows, allowed_types):
        """Output Excel has exactly two sheets: Accounts and Allowed Types."""
        buffer = write_openfloat_excel(sample_output_rows, allowed_types)
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        assert "Accounts" in wb.sheetnames
        assert "Allowed Types" in wb.sheetnames
        assert len(wb.sheetnames) == 2
        wb.close()

    def test_accounts_sheet_columns(self, sample_output_rows, allowed_types):
        """Accounts sheet has the correct 8 columns."""
        buffer = write_openfloat_excel(sample_output_rows, allowed_types)
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Accounts"]
        headers = [cell.value for cell in ws[1]]
        expected = [
            "Account Type",
            "Account Name",
            "Account Number",
            "Till or Paybill Number",
            "Till or Paybill Business Name",
            "Notification Phone Number",
            "Amount",
            "Remark",
        ]
        assert headers == expected
        wb.close()

    def test_accounts_sheet_data(self, sample_output_rows, allowed_types):
        """Accounts sheet contains the correct data rows."""
        buffer = write_openfloat_excel(sample_output_rows, allowed_types)
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Accounts"]
        # Row 2 should have the first data row
        assert ws.cell(row=2, column=1).value == "Safaricom Prepaid"
        assert ws.cell(row=2, column=2).value == "TEST001"
        assert ws.cell(row=2, column=3).value == "254712345678"
        assert ws.cell(row=2, column=7).value == 150.0
        # Row 3 should have the second data row
        assert ws.cell(row=3, column=1).value == "Airtel Prepaid"
        assert ws.cell(row=3, column=2).value == "TEST002"
        wb.close()

    def test_allowed_types_sheet_verbatim(self, sample_output_rows, allowed_types):
        """Allowed Types sheet matches the reference template verbatim."""
        buffer = write_openfloat_excel(sample_output_rows, allowed_types)
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Allowed Types"]
        output_types = []
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0] is not None:
                output_types.append(str(row[0]))
        assert output_types == allowed_types
        wb.close()

    def test_empty_optional_fields(self, sample_output_rows, allowed_types):
        """Till/Paybill fields are empty (None or empty string) in output."""
        buffer = write_openfloat_excel(sample_output_rows, allowed_types)
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Accounts"]
        # Column 4: Till or Paybill Number, Column 5: Till or Paybill Business Name
        # openpyxl may store empty strings as None
        assert ws.cell(row=2, column=4).value in ("", None)
        assert ws.cell(row=2, column=5).value in ("", None)
        wb.close()