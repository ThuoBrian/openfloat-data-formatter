"""OpenFloat Transaction Statement parsing, reporting, and reconciliation.

After staff upload an ADF-formatted file to the OpenFloat SaaS, OpenFloat
produces a 'Transaction Statement' export (one .xlsx per batch). This module
parses those statements, reports on successful vs unsuccessful transactions,
and optionally reconciles the statement against the original Process Maker
input to catch beneficiaries who were never paid.

Statement format facts (verified against real exports):
- Single sheet 'Transaction Statement'; column count VARIES between exports
  (12 or 13 columns — 'Reference Id' present only in some), so parsing is
  header-driven by column NAME, never by position. 'Amount' is always last.
- 'Successful' rows carry an Amount; 'Reversed' rows have Amount=None and a
  'Reference Id' pointing at the original transaction.
- Footer row: every cell empty except Amount = grand total.
- 'Account Number' is stored as an int (254XXXXXXXXX); 'Account Name' is an
  int or a string id like 'I220008' — all cells are str-coerced.
- Date cells are 'DD/MM/YYYY hh:mm:ss AM/PM' strings.
- The Remark column holds the case reference in the raw case_remark format,
  parsed by the existing normalizer.parse_case_remark. The amount embedded in
  the remark is the PER-CASE TOTAL, not a per-row amount.

Classification rule: 'Successful' (case-insensitive) = paid; ANY other status
is unsuccessful and flagged for follow-up. Unknown statuses never crash —
they are counted and classified unsuccessful.
"""

from __future__ import annotations

import math
from collections import Counter, defaultdict
from collections.abc import Sequence
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .config import Settings, settings
from .models import (
    CaseRollup,
    ReconciliationEntry,
    ReconciliationResult,
    StatementFileSummary,
    StatementReport,
    StatementTransaction,
)
from .normalizer import normalize_phone, resolve_case_remark

STATEMENT_SHEET_NAME = "Transaction Statement"

# Columns every statement must have (any position). 'Approval Id',
# 'Transaction Type', 'Initiated By', 'Approved/Rejected By' and
# 'Reference Id' are picked up by name when present but not required.
REQUIRED_STATEMENT_COLUMNS = [
    "Transaction Id",
    "Transaction Status",
    "Date",
    "Account Name",
    "Account Number",
    "Account Type",
    "Remark",
    "Amount",
]

_STATEMENT_DATE_FORMAT = "%d/%m/%Y %I:%M:%S %p"
_STATEMENT_DATE_FALLBACK_FORMATS = ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y")
_EXCEL_SERIAL_EPOCH = datetime(1899, 12, 30)  # Excel day 0 (Windows convention)
_AMOUNT_TOLERANCE = 0.01


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _coerce_cell(raw: Any) -> str:
    """str()-coerce a statement cell for display/comparison.

    Handles None/NaN (→ ''), int floats (27550.0 → '27550'), and strips
    surrounding whitespace from strings.
    """
    if raw is None:
        return ""
    if isinstance(raw, float):
        if math.isnan(raw):
            return ""
        if raw.is_integer():
            return str(int(raw))
        return str(raw)
    return str(raw).strip()


def parse_statement_date(raw: Any) -> tuple[datetime | None, str | None]:
    """Parse a statement Date cell.

    Handles strings ('24/08/2026 02:56:42 PM'), datetimes (openpyxl may
    auto-parse some cells), ints/floats (Excel serial dates), and None.

    Returns:
        A tuple of (datetime, None) on success, (None, error) on a
        non-empty value that can't be parsed, (None, None) when empty.
    """
    if raw is None or (isinstance(raw, float) and math.isnan(raw)):
        return (None, None)
    if isinstance(raw, datetime):
        return (raw, None)
    if isinstance(raw, str):
        raw_str = raw.strip()
        if not raw_str:
            return (None, None)
        for fmt in (_STATEMENT_DATE_FORMAT, *_STATEMENT_DATE_FALLBACK_FORMATS):
            try:
                return (datetime.strptime(raw_str, fmt), None)
            except ValueError:
                continue
        return (None, f"Date '{raw_str}' does not match expected format {_STATEMENT_DATE_FORMAT!r}")
    if isinstance(raw, (int, float)):
        try:
            return (_EXCEL_SERIAL_EPOCH + timedelta(days=float(raw)), None)
        except (ValueError, OverflowError):
            return (None, f"Date serial '{raw}' is out of range")
    return (None, f"Unrecognized date value {raw!r}")


def is_successful_status(status: Any) -> bool:
    """True when a Transaction Status cell means 'paid'.

    Case-insensitive and whitespace-tolerant ('Successful ', 'SUCCESSFUL').
    """
    return _coerce_cell(status).lower() == "successful"


def _is_blank(value: Any) -> bool:
    """True for None, NaN, and empty/whitespace-only strings."""
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and not value.strip()


def _is_footer_row(row_map: dict[str, Any]) -> bool:
    """True for the grand-total footer row: every column empty except Amount."""
    return all(
        _is_blank(value) for name, value in row_map.items() if name != "Amount"
    ) and not _is_blank(row_map.get("Amount"))


def _parse_amount(raw: Any, file_name: str, row_number: int, warnings: list[str]) -> float | None:
    """Coerce an Amount cell; None/blank stays None (Reversed rows), malformed warns."""
    if _is_blank(raw):
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        warnings.append(
            f"{file_name} row {row_number}: Amount {raw!r} is not numeric — recorded as empty"
        )
        return None


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


def _build_transaction(
    row_map: dict[str, Any],
    row_number: int,
    file_name: str,
    country_prefix: str,
    warnings: list[str],
) -> StatementTransaction:
    """Build one StatementTransaction from a header-name → cell map."""
    status = _coerce_cell(row_map.get("Transaction Status"))
    date_raw = row_map.get("Date")
    date, date_error = parse_statement_date(date_raw)
    if date_error is not None:
        warnings.append(f"{file_name} row {row_number}: {date_error}")

    account_raw = row_map.get("Account Number")
    # An empty cell reads as None — normalize_phone would stringify it as 'None'
    account_number, account_error = normalize_phone(
        "" if account_raw is None else account_raw, country_prefix
    )
    account_number_error: str | None = None
    if account_error is not None:
        account_number_error = account_error
        account_number = _coerce_cell(account_raw)
        warnings.append(f"{file_name} row {row_number}: {account_error} — kept raw value")

    remark_raw, remark_parts, remark_error = resolve_case_remark(row_map.get("Remark"))
    if remark_error is not None:
        warnings.append(f"{file_name} row {row_number}: {remark_error}")

    return StatementTransaction(
        file_name=file_name,
        row_number=row_number,
        approval_id=_coerce_cell(row_map.get("Approval Id")),
        transaction_id=_coerce_cell(row_map.get("Transaction Id")),
        transaction_type=_coerce_cell(row_map.get("Transaction Type")),
        status=status,
        is_successful=is_successful_status(status),
        date_raw=_coerce_cell(date_raw),
        date=date,
        account_name=_coerce_cell(row_map.get("Account Name")),
        account_number=account_number,
        account_number_error=account_number_error,
        account_type=_coerce_cell(row_map.get("Account Type")),
        remark=remark_raw,
        remark_parts=remark_parts,
        initiated_by=_coerce_cell(row_map.get("Initiated By")),
        approved_rejected_by=_coerce_cell(row_map.get("Approved/Rejected By")),
        reference_id=_coerce_cell(row_map.get("Reference Id")),
        amount=_parse_amount(row_map.get("Amount"), file_name, row_number, warnings),
    )


def parse_statement_file(
    source: str | Path | BytesIO,
    source_name: str = "",
    country_prefix: str = "254",
) -> tuple[list[StatementTransaction], float | None, list[str], list[str]]:
    """Parse one OpenFloat Transaction Statement export.

    Header-driven: builds a {header_name: column_index} map from row 1, so
    12-column and 13-column exports both parse. Blank rows are skipped; the
    grand-total footer row is captured as footer_total, not a transaction.
    Phone normalization, remark parsing, and date parsing each soft-fail
    with a warning instead of rejecting the row.

    Args:
        source: Path to the .xlsx, or an in-memory BytesIO (file uploader).
        source_name: Display name for warnings/errors (defaults to the file name).
        country_prefix: Country code used to normalize Account Numbers.

    Returns:
        A tuple of (transactions, footer_total, warnings, errors).
        Structural problems (missing sheet/columns) are returned as error
        strings with an empty transaction list; unreadable files raise.
    """
    file_name = source_name or getattr(source, "name", "") or (
        Path(source).name if isinstance(source, (str, Path)) else "statement"
    )
    warnings: list[str] = []
    errors: list[str] = []

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if STATEMENT_SHEET_NAME not in workbook.sheetnames:
            return (
                [],
                None,
                warnings,
                [
                    f"'{file_name}' has no '{STATEMENT_SHEET_NAME}' sheet "
                    f"(found: {', '.join(workbook.sheetnames)})"
                ],
            )
        worksheet = workbook[STATEMENT_SHEET_NAME]

        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return ([], None, warnings, [f"'{file_name}' sheet is empty"])

        header_map = {
            _coerce_cell(cell): index
            for index, cell in enumerate(header_row)
            if not _is_blank(cell)
        }
        missing = [col for col in REQUIRED_STATEMENT_COLUMNS if col not in header_map]
        if missing:
            return (
                [],
                None,
                warnings,
                [f"'{file_name}' is missing required column(s): {', '.join(missing)}"],
            )

        footer_total: float | None = None
        transactions: list[StatementTransaction] = []

        for row_number, row in enumerate(rows_iter, start=2):
            if row is None or all(_is_blank(cell) for cell in row):
                continue
            row_map = {
                name: row[index] if index < len(row) else None
                for name, index in header_map.items()
            }
            if _is_footer_row(row_map):
                footer_total = _parse_amount(
                    row_map.get("Amount"), file_name, row_number, warnings
                )
                continue
            transactions.append(
                _build_transaction(row_map, row_number, file_name, country_prefix, warnings)
            )
    finally:
        workbook.close()

    return transactions, footer_total, warnings, errors


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------


def summarize_transactions(
    transactions: Sequence[StatementTransaction],
    file_name: str = "",
    footer_total: float | None = None,
) -> StatementFileSummary:
    """Aggregate counts, status histogram, and totals for a set of transactions."""
    total_rows = len(transactions)
    successful_count = sum(1 for txn in transactions if txn.is_successful)
    total_disbursed = sum(
        txn.amount for txn in transactions if txn.is_successful and txn.amount is not None
    )
    footer_matches = (
        None if footer_total is None else abs(footer_total - total_disbursed) < _AMOUNT_TOLERANCE
    )
    return StatementFileSummary(
        file_name=file_name,
        total_rows=total_rows,
        successful_count=successful_count,
        unsuccessful_count=total_rows - successful_count,
        counts_by_status=dict(Counter(txn.status for txn in transactions if txn.status)),
        total_disbursed=total_disbursed,
        footer_total=footer_total,
        footer_matches=footer_matches,
        success_rate=(successful_count / total_rows) if total_rows else 0.0,
        unparsed_remarks=sum(
            1 for txn in transactions if txn.remark and txn.remark_parts is None
        ),
        unparsed_dates=sum(1 for txn in transactions if txn.date_raw and txn.date is None),
    )


def rollup_by_case(
    transactions: Sequence[StatementTransaction],
) -> tuple[list[CaseRollup], int]:
    """Group transactions by the case parsed from their Remark.

    Rows whose Remark failed to parse are excluded and counted in the
    returned unparsed count. `difference` is disbursed_total - remark_amount
    (the remark amount is the per-case total), the soft shortfall flag.

    Returns:
        A tuple of (rollups sorted by case_number, unparsed_remark_count).
    """
    groups: dict[tuple[str, str, str], dict[str, Any]] = {}
    unparsed = 0

    for txn in transactions:
        if txn.remark_parts is None:
            if txn.remark:
                unparsed += 1
            continue
        parts = txn.remark_parts
        key = (parts.case_number, parts.project_code, parts.activity_code)
        group = groups.setdefault(
            key,
            {
                "remark_amount": float(parts.amount),
                "total_rows": 0,
                "successful_count": 0,
                "unsuccessful_count": 0,
                "disbursed_total": 0.0,
            },
        )
        group["total_rows"] += 1
        if txn.is_successful:
            group["successful_count"] += 1
            group["disbursed_total"] += txn.amount or 0.0
        else:
            group["unsuccessful_count"] += 1

    rollups = [
        CaseRollup(
            case_number=case_number,
            project_code=project_code,
            activity_code=activity_code,
            remark_amount=group["remark_amount"],
            total_rows=group["total_rows"],
            successful_count=group["successful_count"],
            unsuccessful_count=group["unsuccessful_count"],
            disbursed_total=group["disbursed_total"],
            difference=group["disbursed_total"] - group["remark_amount"],
        )
        for (case_number, project_code, activity_code), group in groups.items()
    ]
    rollups.sort(key=lambda rollup: rollup.case_number)
    return rollups, unparsed


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------


def reconcile(
    input_df: pd.DataFrame,
    transactions: Sequence[StatementTransaction],
    config: Settings | None = None,
) -> ReconciliationResult:
    """Match Process Maker input beneficiaries against statement transactions.

    Match key is the normalized phone (input `airtime_phone` and statement
    `Account Number` both pass through normalize_phone). Input rows with
    unnormalizable phones land in missing_from_statement with the phone error
    in their notes — never silently dropped.

    Classification per input phone:
    - >=1 Successful statement row   → matched_paid
    - only unsuccessful statement rows → matched_not_paid
    - no statement row at all        → missing_from_statement

    Soft amount flag: input amount vs the phone's Successful total → note.
    Duplicates: input phone seen >1 times → duplicate_input_phones;
    statement phone with >1 Successful rows → multiply_paid_phones.
    """
    if config is None:
        config = settings
    prefix = config.default_country_prefix

    # --- Statement side: index transactions by normalized phone ---
    statement_by_phone: dict[str, list[StatementTransaction]] = defaultdict(list)
    for txn in transactions:
        if txn.account_number and txn.account_number_error is None:
            statement_by_phone[txn.account_number].append(txn)

    # --- Input side: normalize phones, gather amounts/row numbers ---
    input_phone_counts: Counter[str] = Counter()
    normalized_rows: list[tuple[int, str, str | None, Any, float | None]] = []
    for idx, row in input_df.iterrows():
        row_number = idx + 2  # 1-based, accounting for the header row
        phone, phone_error = normalize_phone(row.get("airtime_phone", ""), prefix)
        if phone:
            input_phone_counts[phone] += 1
        amount_raw = row.get("amount", None)
        amount: float | None
        try:
            amount = float(amount_raw)
        except (ValueError, TypeError):
            amount = None
        normalized_rows.append((row_number, phone, phone_error, row, amount))

    entries: dict[str, ReconciliationEntry] = {}
    for row_number, phone, phone_error, row, amount in normalized_rows:
        unique_id = str(row.get("unique_id", "")).strip()
        if phone_error is not None:
            # Unnormalizable input phone: keep the raw value as the key and
            # surface the error — it can never match a statement row.
            key = str(row.get("airtime_phone", "")).strip()
            entry = entries.setdefault(
                key,
                ReconciliationEntry(phone=key, unique_id=unique_id),
            )
            if phone_error not in entry.notes:
                entry.notes.append(phone_error)
        else:
            entry = entries.setdefault(
                phone,
                ReconciliationEntry(phone=phone, unique_id=unique_id),
            )
        entry.input_row_numbers.append(row_number)
        if amount is not None:
            entry.input_amount = (entry.input_amount or 0.0) + amount

    # --- Classify each entry and attach statement-side counts/notes ---
    matched_paid: list[ReconciliationEntry] = []
    matched_not_paid: list[ReconciliationEntry] = []
    missing_from_statement: list[ReconciliationEntry] = []
    input_phones = {phone for _, phone, _, _, _ in normalized_rows if phone}

    for entry in entries.values():
        statement_txns = statement_by_phone.get(entry.phone, [])
        entry.successful_count = sum(1 for txn in statement_txns if txn.is_successful)
        entry.unsuccessful_count = len(statement_txns) - entry.successful_count
        entry.successful_total = sum(
            txn.amount or 0.0 for txn in statement_txns if txn.is_successful
        )

        if entry.successful_count > 0:
            if entry.unsuccessful_count > 0:
                entry.notes.append(
                    f"also has {entry.unsuccessful_count} unsuccessful statement row(s)"
                )
            if entry.input_amount is not None and abs(
                entry.input_amount - entry.successful_total
            ) > _AMOUNT_TOLERANCE:
                entry.notes.append(
                    f"input amount {entry.input_amount:g} differs from statement "
                    f"successful total {entry.successful_total:g}"
                )
            matched_paid.append(entry)
        elif entry.unsuccessful_count > 0:
            matched_not_paid.append(entry)
        else:
            missing_from_statement.append(entry)

    # --- Statement phones absent from the input ---
    statement_not_in_input: list[ReconciliationEntry] = []
    for phone in sorted(set(statement_by_phone) - input_phones):
        txns = statement_by_phone[phone]
        statement_not_in_input.append(
            ReconciliationEntry(
                phone=phone,
                successful_count=sum(1 for txn in txns if txn.is_successful),
                unsuccessful_count=sum(1 for txn in txns if not txn.is_successful),
                successful_total=sum(
                    txn.amount or 0.0 for txn in txns if txn.is_successful
                ),
                notes=["not present in the input file"],
            )
        )

    multiply_paid_phones = sorted(
        phone
        for phone, txns in statement_by_phone.items()
        if sum(1 for txn in txns if txn.is_successful) > 1
    )

    return ReconciliationResult(
        input_rows=len(input_df),
        matched_paid=sorted(matched_paid, key=lambda entry: entry.phone),
        matched_not_paid=sorted(matched_not_paid, key=lambda entry: entry.phone),
        missing_from_statement=sorted(missing_from_statement, key=lambda entry: entry.phone),
        statement_not_in_input=statement_not_in_input,
        duplicate_input_phones=sorted(
            phone for phone, count in input_phone_counts.items() if count > 1
        ),
        multiply_paid_phones=multiply_paid_phones,
    )


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


def build_statement_report(
    sources: Sequence[str | Path | BytesIO],
    source_names: Sequence[str] | None = None,
    input_df: pd.DataFrame | None = None,
    config: Settings | None = None,
) -> StatementReport:
    """Build the full statement report across one or more statement exports.

    Parses every source, summarizes each file plus a combined summary, and
    runs reconciliation when a Process Maker input DataFrame is supplied.
    Per-file problems are collected into errors/warnings so one bad upload
    never kills the others' report.

    Args:
        sources: Statement .xlsx paths or in-memory BytesIO buffers.
        source_names: Display names matching `sources` (defaults to file names).
        input_df: Optional Process Maker input DataFrame for reconciliation.
        config: Optional settings override. Uses global defaults if None.
    """
    if config is None:
        config = settings
    if source_names is None:
        source_names = [
            Path(source).name if isinstance(source, (str, Path)) else f"statement {index + 1}"
            for index, source in enumerate(sources)
        ]

    warnings: list[str] = []
    errors: list[str] = []
    transactions: list[StatementTransaction] = []
    file_summaries: list[StatementFileSummary] = []

    for source, name in zip(sources, source_names, strict=True):
        try:
            file_transactions, footer_total, file_warnings, file_errors = parse_statement_file(
                source, source_name=name, country_prefix=config.default_country_prefix
            )
        except Exception as exc:  # unreadable/corrupt file — keep other files' reports
            errors.append(f"'{name}' could not be read: {exc}")
            continue
        warnings.extend(file_warnings)
        errors.extend(file_errors)
        transactions.extend(file_transactions)
        file_summaries.append(summarize_transactions(file_transactions, name, footer_total))

    rollups, _unparsed = rollup_by_case(transactions)

    return StatementReport(
        file_summaries=file_summaries,
        combined=summarize_transactions(transactions, "Combined"),
        transactions=transactions,
        unsuccessful_transactions=[txn for txn in transactions if not txn.is_successful],
        case_rollups=rollups,
        warnings=warnings,
        errors=errors,
        reconciliation=(
            reconcile(input_df, transactions, config) if input_df is not None else None
        ),
    )
