"""Generate the Process Maker input template.

Produces `docs/processmaker-input-template.xlsx`: a ready-to-fill export
template for staff, with the exact columns the formatter expects, a couple
of valid example rows, in-cell dropdown validation for the two enum-like
columns (`consent`, `network`), and an `Instructions` sheet documenting the
per-column format rules (see CLAUDE.md "Key Domain Rules").

Column order is imported from `openfloat_formatter.config` so this template can
never drift from the schema the pipeline actually validates against.

Usage (repo root, after `uv sync`):
    uv run python scripts/generate_processmaker_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from copy import copy

from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from openfloat_formatter.config import DEFAULT_NETWORK_MAP, PROCESSMAKER_COLUMNS

PROJECT_ROOT = Path(__file__).resolve().parent.parent

OUTPUT_PATH = PROJECT_ROOT / "docs" / "processmaker-input-template.xlsx"

# The 12th column: optional, not part of PROCESSMAKER_COLUMNS, but read by
# the transformer when present.
CASE_REMARK_COLUMN = "case_remark"
TEMPLATE_COLUMNS = [*PROCESSMAKER_COLUMNS, CASE_REMARK_COLUMN]

CONSENT_CHOICES = ["Yes", "No"]
NETWORK_CHOICES = list(DEFAULT_NETWORK_MAP.keys())

# Sample rows demonstrating valid data, mirroring src/tests/conftest.py::minimal_df
SAMPLE_ROWS = [
    {
        "unique_id": "TEST001",
        "consent": "Yes",
        "airtime_phone": "712345678",
        "network": "Safaricom",
        "submissiondate": "8/25/2026 10:00",
        "today": "25aug2026",
        "amount": 150,
        "project_name": "Test Project",
        "Project_Activity": "g05|Testing",
        "department": "Projects",
        "survey": "Baseline",
        "case_remark": "C#37166 22505AA RESP AIRTIME-KSH150 d05",
    },
    {
        "unique_id": "TEST002",
        "consent": "Yes",
        "airtime_phone": "798765432",
        "network": "Airtel",
        "submissiondate": "8/25/2026 11:00",
        "today": "25aug2026",
        "amount": 200,
        "project_name": "Test Project",
        "Project_Activity": "g05|Testing",
        "department": "Projects",
        "survey": "Baseline",
        "case_remark": "C#37167 22505AA RESP AIRTIME-KSH200 d05",
    },
    {
        "unique_id": "TEST003",
        "consent": "Yes",
        "airtime_phone": "722334455",
        "network": "Telkom",
        "submissiondate": "8/25/2026 12:00",
        "today": "25aug2026",
        "amount": 100,
        "project_name": "Test Project",
        "Project_Activity": "g05|Testing",
        "department": "Projects",
        "survey": "Baseline",
        # case_remark left blank on purpose -> demonstrates the legacy
        # "{project_name} - {Project_Activity}" fallback.
        "case_remark": "",
    },
]

INSTRUCTIONS = [
    ("unique_id", "Optional. Any identifier; written verbatim to the output 'Account Name' column."),
    ("consent", 'Required. Must be exactly "Yes" (case-insensitive). Any other value excludes the row.'),
    (
        "airtime_phone",
        "Required. Kenyan phone number. Digits only after cleanup; a leading 254 or 0 prefix is stripped "
        "(0 only if 9 digits would remain longer than 9); must resolve to exactly 9 digits. "
        "Example: 254712345678 or 0712345678 or 712345678 all normalize the same way.",
    ),
    (
        "network",
        "Required, case-sensitive. Must be exactly one of: " + ", ".join(NETWORK_CHOICES) + ". Any other value is a hard error.",
    ),
    ("submissiondate", "Optional. Not validated by the formatter; informational only."),
    ("today", "Optional. Not validated by the formatter; informational only."),
    (
        "amount",
        "Required. A positive number (KES). Zero, negative, or non-numeric values are rejected. "
        "Amounts above KES 10,000 are allowed but flagged as a warning.",
    ),
    ("project_name", "Optional. Used only as part of the legacy Remark fallback when case_remark is blank."),
    ("Project_Activity", "Optional (note capitalization). Used only as part of the legacy Remark fallback when case_remark is blank."),
    ("department", "Optional. Not validated by the formatter; informational only."),
    ("survey", "Optional. Not validated by the formatter; informational only."),
    (
        "case_remark",
        "Optional but recommended. Fixed format: 'C#<case_number> <project_code> RESP AIRTIME-KSH<amount> <activity_code>', "
        "e.g. 'C#37166 22505AA RESP AIRTIME-KSH29400 d05'. Reformatted automatically to "
        "'Case #<case_number> | <project_code> | RESP | AIRTIME KSH <amount> | <activity_code>'. "
        "If left blank, the Remark falls back to '{project_name} - {Project_Activity}'. "
        "The embedded amount is cross-checked against the row's real amount column (mismatch is a soft warning only).",
    ),
]

DROPDOWN_ROWS = 500  # how many data rows below the header get dropdown validation


def build_template_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Template"

    header_font = Font(bold=True)
    for col_idx, column_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(14, len(column_name) + 2)

    for row_idx, sample in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, column_name in enumerate(TEMPLATE_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sample.get(column_name, ""))

    last_row = DROPDOWN_ROWS + 1

    consent_col = TEMPLATE_COLUMNS.index("consent") + 1
    consent_letter = ws.cell(row=1, column=consent_col).column_letter
    consent_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(CONSENT_CHOICES)}"',
        allow_blank=True,
        showDropDown=False,
    )
    consent_dv.error = 'Consent must be "Yes" or "No".'
    consent_dv.errorTitle = "Invalid consent"
    ws.add_data_validation(consent_dv)
    consent_dv.add(f"{consent_letter}2:{consent_letter}{last_row}")

    network_col = TEMPLATE_COLUMNS.index("network") + 1
    network_letter = ws.cell(row=1, column=network_col).column_letter
    network_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(NETWORK_CHOICES)}"',
        allow_blank=True,
        showDropDown=False,
    )
    network_dv.error = "Network must exactly match one of the allowed values (case-sensitive)."
    network_dv.errorTitle = "Invalid network"
    ws.add_data_validation(network_dv)
    network_dv.add(f"{network_letter}2:{network_letter}{last_row}")

    ws.freeze_panes = "A2"


def build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    header_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Column").font = header_font
    ws.cell(row=1, column=2, value="Rule").font = header_font
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 110

    for row_idx, (column_name, rule) in enumerate(INSTRUCTIONS, start=2):
        ws.cell(row=row_idx, column=1, value=column_name)
        cell = ws.cell(row=row_idx, column=2, value=rule)
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        cell.alignment = alignment

    ws.freeze_panes = "A2"


def main() -> None:
    wb = Workbook()
    build_template_sheet(wb)
    build_instructions_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote template to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
